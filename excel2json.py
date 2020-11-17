#!/usr/bin/env python3
import json
import sys
from urllib.request import urlretrieve

import pandas as pd
from openpyxl import load_workbook

URL = 'https://dshs.texas.gov/coronavirus/TexasCOVID19CaseCountData.xlsx'
BASENAME = 'TexasCOVID19CaseCountData'

# Read the latest population estimate from Texas Demographic Center https://demographics.texas.gov
# urlretrieve('https://demographics.texas.gov/Resources/TPEPP/Estimates/2018/2018_txpopest_county.csv',
#             'Texas2018PopulationEstimate.csv')
county_population_df = pd.read_csv('Texas2018PopulationEstimate.csv', encoding="utf-8")
county_population_df.loc[county_population_df['county'] == 'De Witt', 'county'] = 'DeWitt'
county_population_df = county_population_df[['county', 'jan1_2019_pop_est']]
county_population_df = county_population_df.rename(columns={"jan1_2019_pop_est": "population"})
county_population_df = county_population_df.set_index('county')

try:
    urlretrieve(URL, BASENAME + '.xlsx')
except Exception as e:
    print(e)
    sys.exit(-1)
else:
    wb = load_workbook(filename=BASENAME + '.xlsx')
    ws = wb['Case and Fatalities']

    TXCases = {
        "date": None,
        "hospitalizations": None,
        "positivity rate": None,
        "counts": []
    }

    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        TXCases["date"] = row[0]
        print(TXCases["date"])

    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] in county_population_df.index:
            entry = {
                "county": row[0],
                "population": int(county_population_df.loc[row[0]].population),
                "cases": row[1],
                "fatalities": row[2]
            }
        elif row[0] == "Total":
            entry = {
                "county": row[0],
                "population": int(county_population_df.loc['State of Texas'].population),
                "cases": row[1],
                "fatalities": row[2]
            }
        else:
            continue

        TXCases["counts"].append(entry)

    ws = wb['Hospitalizations']
    for row in ws.iter_rows(min_row=3, max_row=3, values_only=True):
        TXCases["hospitalizations"] = row[1]

    if str(TXCases["hospitalizations"]).lower() == 'count':
        for row in ws.iter_rows(min_row=4, max_row=4, values_only=True):
            TXCases["hospitalizations"] = row[1]

    print("hospitalizations:", TXCases["hospitalizations"])

    ws = wb['Tests by Day']
    for row in ws.iter_rows(min_row=4, max_row=4, values_only=True):
        TXCases["positivity rate"] = row[3]
        print("Positivity Rate:", TXCases["positivity rate"])

    with open(BASENAME + '.json', 'w', encoding='utf-8') as jsonFile:
        jsonFile.write(json.dumps(TXCases))
